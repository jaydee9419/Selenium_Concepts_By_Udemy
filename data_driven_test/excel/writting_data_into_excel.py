import openpyxl

file_path = r'C:\\Users\\11571\\Desktop\\new.xlsx'

workbook = openpyxl.load_workbook(file_path)

sheet = workbook.create_sheet("new")
# sheet = workbook.active # if we have only one sheet in a file




for r in range(1, 6):
    for c in range(1,6):
        if r == 2 and c in range(1, 6):
            sheet.cell(2, c).value='change'
        else:
            sheet.cell(r, c).value='B'

workbook.save(file_path)


