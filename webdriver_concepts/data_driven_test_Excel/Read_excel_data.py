import openpyxl

# path of the file
file = "C:\\Users\\11571\\Downloads\\data.xlsx"

#Loading the excel file
workbook = openpyxl.load_workbook(file)

#inside excel accessing the required sheet
sheet = workbook['Sheet1']

#read the count of rows
rows = sheet.max_row
print(rows)

#read the count of columns
columns = sheet.max_column
print(columns)

#read all the rows and columns from the excel sheet
for row in range(1, rows+1):
    for column in range(1, columns+1):
        print(sheet.cell(row,column).value, end="         ")
    print()
   
